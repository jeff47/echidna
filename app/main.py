from __future__ import annotations

import html
import importlib.metadata as importlib_metadata
import io
import json
import logging
import os
import re
import secrets
import subprocess
import threading
import time
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, cast

from fastapi import Depends, FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, Response
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.templating import Jinja2Templates

from app.logic import (
    AnalysisResult,
    analyze_citations,
    build_clusters,
    citation_in_window,
    citation_matches_excluded_type,
    default_excluded_type_terms,
    format_summary,
    parse_type_terms,
    parse_pmids,
    parse_target_name,
    parse_target_orcid,
    _institution_key,
    _has_distinctive_institution_overlap,
)
from app.auth import verify_password_against_hash
from app.models import AuthorMatch, Citation, Cluster
from app.ncbi import NcbiClient, _normalize_affiliation_text, split_chunks
from app.orcid import OrcidClient
from app.usage import fetch_usage_run_by_id, fetch_usage_runs, record_usage_run, update_usage_run

logging.basicConfig(
    level=os.getenv("ECHIDNA_LOG_LEVEL", "INFO").upper(),
    format="%(asctime)s %(levelname)s %(name)s %(message)s",
)
logger = logging.getLogger(__name__)

app = FastAPI(title="ECHIDNA")
templates = Jinja2Templates(directory="templates")
client = NcbiClient()
orcid_client = OrcidClient()
RUNS_DIR = Path(os.getenv("ECHIDNA_RUNS_DIR", "/tmp/echidna-runs"))
USAGE_DB_PATH = Path(os.getenv("ECHIDNA_USAGE_DB_PATH", str(RUNS_DIR.parent / "usage.db")))
admin_basic = HTTPBasic(auto_error=False)


@dataclass(slots=True)
class RunState:
    author_name: str
    target_orcid: str
    start_year: int | None
    end_year: int | None
    excluded_type_terms: list[str]
    citations: list[Citation]
    clusters: list[Cluster]
    matches: list[AuthorMatch]
    excluded_pmids: set[str]
    excluded_reasons: dict[str, str]
    orcid_identifier_matches: dict[str, list[str]]
    orcid_affiliation_matches: dict[str, dict[str, str]]
    orcid_sync_error: str | None
    created_at: float = field(default_factory=time.time)


RUNS: dict[str, RunState] = {}
RUN_ID_PATTERN = re.compile(r"^[a-f0-9]{32}$")
YEAR_OPTION_MIN = 1950
MAX_UPLOAD_BYTES = 64 * 1024
MAX_PMID_COUNT = 500
PROJECT_ROOT = Path(__file__).resolve().parent.parent
LANDING_IMAGE_PATH = PROJECT_ROOT / "echidna.jpg"
FAVICON_PATH = PROJECT_ROOT / "favicon.ico"
ORCID_AFFILIATION_LEVEL_RANK = {"verified": 4, "likely": 3, "contains": 2, "possible": 1}
CITATION_SOURCE_ORDER = ("user", "litsense")
PREPRINT_PUBLICATION_TYPE = "preprint"
CITATION_SOURCE_BADGE_META: dict[str, dict[str, str]] = {
    "user": {
        "label": "user",
        "class": "source-badge source-user",
        "title": "Citation added from user file",
    },
    "litsense": {
        "label": "litsense",
        "class": "source-badge source-litsense",
        "title": "Citation added from LitSense",
    },
}


def _positive_int_env(name: str, default: int) -> int:
    raw = os.getenv(name, "").strip()
    if not raw:
        return default
    try:
        value = int(raw)
    except ValueError:
        logger.warning("Invalid integer for %s=%r; using default %d", name, raw, default)
        return default
    return max(1, value)


MAX_CONCURRENT_DISAMBIGUATIONS = _positive_int_env("ECHIDNA_MAX_CONCURRENT_DISAMBIGUATIONS", 1)
BUSY_RETRY_AFTER_SECONDS = _positive_int_env("ECHIDNA_BUSY_RETRY_AFTER_SECONDS", 30)
RUN_TTL_SECONDS = _positive_int_env("ECHIDNA_RUN_TTL_SECONDS", 24 * 60 * 60)
DISAMBIGUATION_SEMAPHORE = threading.BoundedSemaphore(MAX_CONCURRENT_DISAMBIGUATIONS)


def _short_commit(commit: str | None, *, length: int = 12) -> str | None:
    if not commit:
        return None
    return commit[:length]


def _distribution_version(dist_name: str) -> str | None:
    try:
        return importlib_metadata.version(dist_name)
    except importlib_metadata.PackageNotFoundError:
        return None


def _distribution_vcs_commit(dist_name: str) -> str | None:
    try:
        dist = importlib_metadata.distribution(dist_name)
    except importlib_metadata.PackageNotFoundError:
        return None
    direct_url = dist.read_text("direct_url.json")
    if not direct_url:
        return None
    try:
        payload = json.loads(direct_url)
    except json.JSONDecodeError:
        return None
    vcs_info = payload.get("vcs_info")
    if not isinstance(vcs_info, dict):
        return None
    commit = vcs_info.get("commit_id")
    if not isinstance(commit, str) or not commit:
        return None
    return commit


def _git_commit(repo_root: Path) -> str | None:
    try:
        result = subprocess.run(
            ["git", "-C", str(repo_root), "rev-parse", "HEAD"],
            check=True,
            capture_output=True,
            text=True,
            timeout=2,
        )
    except (FileNotFoundError, subprocess.SubprocessError):
        return None
    commit = result.stdout.strip()
    return commit or None


def _version_label(version: str | None, commit: str | None) -> str:
    return f"{version or 'unknown'} ({_short_commit(commit) or 'unknown'})"


def _build_footer_metadata() -> dict[str, str]:
    build_datetime = os.getenv("ECHIDNA_BUILD_DATETIME", "").strip()
    if not build_datetime:
        build_datetime = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    app_version = _distribution_version("echidna")
    app_commit = os.getenv("ECHIDNA_APP_COMMIT", "").strip() or _git_commit(PROJECT_ROOT)
    normalizer_version = _distribution_version("affiliation_normalizer")
    normalizer_commit = _distribution_vcs_commit("affiliation_normalizer")
    return {
        "build_datetime": build_datetime,
        "app_label": _version_label(app_version, app_commit),
        "normalizer_label": _version_label(normalizer_version, normalizer_commit),
    }


templates.env.globals["build_info"] = _build_footer_metadata()


def _to_author_list(citation: Citation) -> str:
    return ", ".join(f"{a.last_name} {a.initials}".strip() for a in citation.authors)


def _to_author_list_html(citation: Citation, highlight_positions: set[int]) -> str:
    parts: list[str] = []
    for author in citation.authors:
        name = f"{author.last_name} {author.initials}".strip()
        escaped = html.escape(name)
        if author.position in highlight_positions:
            parts.append(f"<strong>{escaped}</strong>")
        else:
            parts.append(escaped)
    return ", ".join(parts)


def _year_options() -> list[int]:
    current = date.today().year
    return list(range(current, YEAR_OPTION_MIN - 1, -1))


def _year_window_label(start_year: int | None, end_year: int | None) -> str:
    if start_year is None or end_year is None:
        return "all years"
    return f"{start_year}-{end_year}"


def _publication_type_exclusion_options() -> list[str]:
    return [term for term in default_excluded_type_terms() if term != PREPRINT_PUBLICATION_TYPE]


def _selected_default_excluded_type_terms(submitted: list[str], options: list[str]) -> list[str]:
    selected = set(parse_type_terms("\n".join(submitted)))
    return [term for term in options if term in selected]


def _dedupe_terms(terms: list[str]) -> list[str]:
    seen: set[str] = set()
    deduped: list[str] = []
    for term in terms:
        normalized = term.strip().lower()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        deduped.append(normalized)
    return deduped


def _dict_row_sort_key(row: dict[str, object]) -> tuple[bool, int, str]:
    year_value = row.get("year")
    year = year_value if isinstance(year_value, int) else None
    pmid = str(row.get("pmid", ""))
    return (year is None, -(year or 0), pmid)

def _orcid_affiliation_level_rank(level: str) -> int:
    return ORCID_AFFILIATION_LEVEL_RANK.get(level.strip().lower(), 0)


def _orcid_badge_class(*, has_identifier_match: bool, affiliation_level: str) -> str:
    if has_identifier_match:
        return "orcid-badge"
    if affiliation_level.strip().lower() == "contains":
        return "orcid-badge orcid-badge-weak"
    return "orcid-badge"


def _orcid_badge_title(
    *,
    has_identifier_match: bool,
    identifier_label: str,
    affiliation_level: str,
    affiliation_label: str,
) -> str:
    if has_identifier_match:
        if identifier_label:
            return f"Citation added from ORCiD record (matched by {identifier_label})"
        return "Citation added from ORCiD record"
    if affiliation_level and affiliation_label:
        return f"ORCiD affiliation {affiliation_level}: {affiliation_label}"
    if affiliation_level:
        return f"ORCiD affiliation {affiliation_level} match"
    return "ORCiD match"


def _citation_source_badges(citation: Citation) -> list[dict[str, str]]:
    badges: list[dict[str, str]] = []
    tags = set(citation.source_tags)
    for tag in CITATION_SOURCE_ORDER:
        if tag not in tags:
            continue
        meta = CITATION_SOURCE_BADGE_META.get(tag)
        if meta is None:
            continue
        badges.append(meta)
    return badges


def _orcid_row_fields(
    *,
    pmid: str,
    target_orcid: str,
    orcid_identifier_matches: dict[str, list[str]] | None,
    orcid_affiliation_matches: dict[str, dict[str, str]] | None,
) -> dict[str, object]:
    identifier_matches = (orcid_identifier_matches or {}).get(pmid, [])
    affiliation_match = (orcid_affiliation_matches or {}).get(pmid, {})
    affiliation_level = str(affiliation_match.get("level", "")).strip().lower()
    affiliation_label = str(affiliation_match.get("label", "")).strip()
    affiliation_institution = str(affiliation_match.get("institution", "")).strip()

    has_identifier_match = bool(identifier_matches)
    has_affiliation_match = bool(affiliation_level and affiliation_label)
    identifier_label = "/".join(value.upper() for value in identifier_matches)

    return {
        "orcid_verified": has_identifier_match,
        "orcid_match_types": list(identifier_matches),
        "orcid_match_label": identifier_label,
        "orcid_affiliation_verified": has_affiliation_match,
        "orcid_affiliation_level": affiliation_level,
        "orcid_affiliation_label": affiliation_label,
        "orcid_affiliation_institution": affiliation_institution,
        # Citation-level iD badge only appears for ORCiD work identifier matches.
        "orcid_badge_show": has_identifier_match,
        "orcid_badge_class": _orcid_badge_class(
            has_identifier_match=has_identifier_match,
            affiliation_level="",
        ),
        "orcid_badge_title": _orcid_badge_title(
            has_identifier_match=has_identifier_match,
            identifier_label=identifier_label,
            affiliation_level=affiliation_level,
            affiliation_label=affiliation_label,
        ),
        "orcid_record_url": f"https://orcid.org/{target_orcid}" if (target_orcid and has_identifier_match) else None,
    }

def _default_correction_mode(row: Any) -> str:
    if not row.include:
        return "exclude"
    if row.is_preprint:
        return "first_senior_preprint" if (row.counted_first or row.counted_senior) else "coauthor_preprint"
    if row.is_review and row.counted_review_senior:
        return "first_senior_review"
    if row.is_review:
        return "coauthor_review"
    if row.counted_first or row.counted_senior:
        return "first_senior_author"
    if row.counted_overall:
        return "coauthor"
    return "exclude"


def _counted_as_category(row: Any) -> str:
    if row.is_preprint:
        return "1st/Sr preprint" if (row.counted_first or row.counted_senior) else "coauthor preprint"
    if row.is_review:
        return "1st/Sr review" if (row.counted_senior or row.counted_review_senior) else "coauthor review"
    return "1st/Sr author" if (row.counted_first or row.counted_senior) else "coauthor"


def _counted_mode_from_flags(row: Any) -> str:
    if row.is_preprint:
        return "first_senior_preprint" if (row.counted_first or row.counted_senior) else "coauthor_preprint"
    if row.is_review:
        return "first_senior_review" if (row.counted_senior or row.counted_review_senior) else "coauthor_review"
    return "first_senior_author" if (row.counted_first or row.counted_senior) else "coauthor"


def _apply_correction_mode(row: Any, mode: str) -> None:
    mode_aliases = {
        "overall": "coauthor",
        "senior": "first_senior_author",
        "review": "coauthor_review",
        "review_senior": "first_senior_review",
    }
    mode = mode_aliases.get(mode, mode)
    if mode == "exclude":
        row.include = False
        row.counted_overall = False
        row.counted_first = False
        row.counted_senior = False
        row.counted_review_senior = False
        row.is_review = False
        row.is_preprint = False
        row.preprint_superseded_by = []
        return
    if mode == "coauthor":
        row.include = True
        row.counted_overall = True
        row.counted_first = False
        row.counted_senior = False
        row.counted_review_senior = False
        row.is_review = False
        row.is_preprint = False
        row.preprint_superseded_by = []
        return
    if mode == "first_senior_author":
        row.include = True
        row.counted_overall = True
        row.counted_first = False
        row.counted_senior = True
        row.counted_review_senior = False
        row.is_review = False
        row.is_preprint = False
        row.preprint_superseded_by = []
        return
    if mode == "coauthor_preprint":
        row.include = True
        row.counted_overall = False
        row.counted_first = False
        row.counted_senior = False
        row.counted_review_senior = False
        row.is_review = False
        row.is_preprint = True
        row.preprint_superseded_by = []
        return
    if mode == "first_senior_preprint":
        row.include = True
        row.counted_overall = False
        row.counted_first = False
        row.counted_senior = True
        row.counted_review_senior = False
        row.is_review = False
        row.is_preprint = True
        row.preprint_superseded_by = []
        return
    if mode == "coauthor_review":
        row.include = True
        row.counted_overall = False
        row.counted_first = False
        row.counted_senior = False
        row.counted_review_senior = False
        row.is_review = True
        row.is_preprint = False
        row.preprint_superseded_by = []
        return
    if mode == "first_senior_review":
        row.include = True
        row.counted_overall = False
        row.counted_first = False
        row.counted_senior = True
        row.counted_review_senior = True
        row.is_review = True
        row.is_preprint = False
        row.preprint_superseded_by = []
        return


def _orcid_identifier_crosscheck_available(*, target_orcid: str, orcid_sync_error: str | None) -> bool:
    if not target_orcid:
        return False
    if not orcid_sync_error:
        return True
    return "identifier cross-check unavailable" not in orcid_sync_error.lower()


def _with_row_render_fields(
    analysis: AnalysisResult,
    *,
    orcid_identifier_matches: dict[str, list[str]] | None = None,
    orcid_affiliation_matches: dict[str, dict[str, str]] | None = None,
    target_orcid: str = "",
    orcid_identifier_checked: bool = False,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for idx, row in enumerate(analysis.rows):
        counted_as = _counted_as_category(row)
        counted_as_report = counted_as if row.include else "exclude"
        counted_mode = _counted_mode_from_flags(row)
        notes = list(row.citation.notes) + list(row.uncertainty_reasons)
        if row.preprint_superseded_by:
            notes.append(
                "Preprint skipped: superseded by peer-reviewed PMID(s): "
                + ", ".join(str(pmid) for pmid in row.preprint_superseded_by)
            )
        if row.forced_include:
            notes.append("Forced include override from Step 1")

        orcid_fields = _orcid_row_fields(
            pmid=row.citation.pmid,
            target_orcid=target_orcid,
            orcid_identifier_matches=orcid_identifier_matches,
            orcid_affiliation_matches=orcid_affiliation_matches,
        )
        if target_orcid and orcid_identifier_checked and not bool(orcid_fields.get("orcid_verified", False)):
            notes.append("No match from ORCiD")

        deduped_notes: list[str] = []
        seen_notes: set[str] = set()
        for note in notes:
            cleaned = str(note).strip()
            if not cleaned or cleaned in seen_notes:
                continue
            seen_notes.add(cleaned)
            deduped_notes.append(cleaned)

        rows.append(
            {
                "idx": idx,
                "pmid": row.citation.pmid,
                "pmcid": row.citation.pmcid,
                "doi": row.citation.doi,
                "pmid_link": f"https://pubmed.ncbi.nlm.nih.gov/{row.citation.pmid}/",
                "pmcid_link": f"https://pmc.ncbi.nlm.nih.gov/articles/{row.citation.pmcid}/" if row.citation.pmcid else None,
                "doi_link": f"https://doi.org/{row.citation.doi}" if row.citation.doi else None,
                **orcid_fields,
                "title": row.citation.title,
                "journal": row.citation.journal,
                "print_date": row.citation.print_date,
                "authors": _to_author_list(row.citation),
                "authors_html": _to_author_list_html(row.citation, row.matched_positions),
                "source_badges": _citation_source_badges(row.citation),
                "publication_types": ", ".join(row.citation.publication_types) if row.citation.publication_types else "(unknown)",
                "counted_as": counted_as,
                "counted_as_report": counted_as_report,
                "correction_default": _default_correction_mode(row),
                "correction_default_uncertain": counted_mode,
                "role_source": row.citation.source_for_roles,
                "notes": "; ".join(deduped_notes),
                "include": row.include,
                "status_label": (
                    "included (review separate)"
                    if row.include and not row.counted_overall and row.is_review
                    else (
                        "included (preprint separate)"
                        if row.include and not row.counted_overall and row.is_preprint
                        else ("included" if row.include else "excluded")
                    )
                ),
            }
        )
    return rows


def _parse_correction_entries(correction_entries: list[str]) -> dict[str, str]:
    corrections_by_pmid: dict[str, str] = {}
    for entry in correction_entries:
        if "::" not in entry:
            continue
        pmid, mode = entry.split("::", 1)
        pmid = pmid.strip()
        mode = mode.strip()
        if not pmid:
            continue
        corrections_by_pmid[pmid] = mode
    return corrections_by_pmid


def _apply_uncertain_selection_and_corrections(
    *,
    analysis: AnalysisResult,
    accepted_row_indices: set[int],
    correction_entries: list[str],
) -> None:
    corrections_by_pmid = _parse_correction_entries(correction_entries)
    for idx in analysis.uncertain_indices:
        row = analysis.rows[idx]
        row.include = idx in accepted_row_indices
        if not row.include:
            continue
        mode = corrections_by_pmid.get(row.citation.pmid)
        if not mode:
            continue
        _apply_correction_mode(row, mode)


def _rebuild_analysis_with_overrides(
    *,
    run: RunState,
    selected_cluster_ids: set[str],
    selected_pmids: set[str],
    selected_excluded_pmids: set[str],
    pmid_filter_enabled: int,
    accepted_uncertain_pmids: set[str],
    correction_entries: list[str],
) -> tuple[AnalysisResult, list[AuthorMatch]]:
    filtered_matches = _filter_matches_for_analysis(
        run=run,
        selected_cluster_ids=selected_cluster_ids,
        selected_pmids=selected_pmids,
        pmid_filter_enabled=pmid_filter_enabled,
        forced_excluded_pmids=selected_excluded_pmids,
    )
    analysis = analyze_citations(
        citations=run.citations,
        matches=filtered_matches,
        selected_cluster_ids=selected_cluster_ids,
        start_year=run.start_year,
        end_year=run.end_year,
        forced_include_pmids=selected_excluded_pmids,
        excluded_pmids=run.excluded_pmids,
    )

    for row in analysis.rows:
        if row.uncertainty_reasons:
            row.include = row.citation.pmid in accepted_uncertain_pmids

    corrections_by_pmid = _parse_correction_entries(correction_entries)
    for row in analysis.rows:
        mode = corrections_by_pmid.get(row.citation.pmid)
        if not mode:
            continue
        _apply_correction_mode(row, mode)

    return analysis, filtered_matches


def _xlsx_response_for_analysis(*, run: RunState, analysis: AnalysisResult) -> Response:
    try:
        from openpyxl import Workbook
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Excel export dependency missing: {exc}") from exc

    row_data = _with_row_render_fields(
        analysis,
        orcid_identifier_matches=run.orcid_identifier_matches,
        orcid_affiliation_matches=run.orcid_affiliation_matches,
        target_orcid=run.target_orcid,
        orcid_identifier_checked=_orcid_identifier_crosscheck_available(
            target_orcid=run.target_orcid,
            orcid_sync_error=run.orcid_sync_error,
        ),
    )
    excluded_rows = [row for row in row_data if not row["include"]]
    included_pubmed_rows = [row for row in row_data if row["include"] and row["role_source"] != "pmc"]
    included_pmc_rows = [row for row in row_data if row["include"] and row["role_source"] == "pmc"]
    report_rows = excluded_rows + included_pubmed_rows + included_pmc_rows

    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise HTTPException(status_code=500, detail="Failed to initialize worksheet for Excel export")
    sheet.title = "Publications"
    sheet.append(
        [
            "PMID/PMCID",
            "PubMed Link",
            "PMC Link",
            "DOI",
            "Title",
            "Journal",
            "Date",
            "Authors",
            "Publication Type",
            "Counted as",
        ]
    )

    for row in report_rows:
        pmid_pmcid = f"PMID {row['pmid']}"
        if row["pmcid"]:
            pmid_pmcid = f"{pmid_pmcid} | {row['pmcid']}"
        sheet.append(
            [
                pmid_pmcid,
                "PubMed",
                "PMC" if row["pmcid_link"] else "",
                row["doi"] or "",
                row["title"],
                row["journal"],
                row["print_date"],
                row["authors"],
                row["publication_types"],
                row["counted_as_report"],
            ]
        )
        current_row = sheet.max_row
        pubmed_cell = sheet.cell(row=current_row, column=2)
        pubmed_cell.hyperlink = str(row["pmid_link"])
        pubmed_cell.style = "Hyperlink"
        if row["pmcid_link"]:
            pmc_cell = sheet.cell(row=current_row, column=3)
            pmc_cell.hyperlink = str(row["pmcid_link"])
            pmc_cell.style = "Hyperlink"
        if row["doi_link"]:
            doi_cell = sheet.cell(row=current_row, column=4)
            doi_cell.hyperlink = str(row["doi_link"])
            doi_cell.style = "Hyperlink"

    sheet.freeze_panes = "A2"
    column_widths = {
        "A": 22,
        "B": 14,
        "C": 10,
        "D": 28,
        "E": 85,
        "F": 24,
        "G": 14,
        "H": 56,
        "I": 24,
        "J": 20,
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    filename = _xlsx_download_filename(run)
    output = io.BytesIO()
    workbook.save(output)
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _filename_author_token(author_name: str) -> str:
    raw = author_name.strip()
    if not raw:
        return "author"
    try:
        target = parse_target_name(raw)
        surname = re.sub(r"[^a-z0-9]+", "", target.surname.lower())
        initials = re.sub(r"[^a-z0-9]+", "", target.initials.lower())
        combined = f"{surname}{initials}"
        if combined:
            return combined
    except ValueError:
        pass
    fallback = re.sub(r"[^a-z0-9]+", "", raw.lower())
    return fallback or "author"


def _xlsx_download_filename(run: RunState) -> str:
    author_token = _filename_author_token(run.author_name)
    if run.start_year is None or run.end_year is None:
        return f"{author_token}_publications_all_years.xlsx"
    return f"{author_token}_publications_{run.start_year}_{run.end_year}.xlsx"


def _missing_affiliation_reason(citation: Citation, affiliation: str) -> str:
    if affiliation.strip():
        return ""

    if not citation.pmcid:
        return "No PMCID mapping, and PubMed author entry has no affiliation."

    if citation.source_for_roles == "pubmed":
        for note in citation.notes:
            if note.startswith("PMC fetch/parse failed"):
                return f"{note}; PubMed author entry has no affiliation."
        return "PMC unavailable, and PubMed author entry has no affiliation."

    if any("Author metadata sourced from PMC" in note for note in citation.notes):
        return "PMC author entry for this contributor does not include affiliation text."

    if any("PMC author metadata incomplete" in note for note in citation.notes):
        return "PMC author metadata incomplete and PubMed fallback also has no affiliation."

    return "Affiliation missing in parsed PMC/PubMed author metadata."


def _build_citation_selection_rows(
    *,
    citations: list[Citation],
    matches: list[AuthorMatch],
    all_matches: list[AuthorMatch] | None = None,
    cluster_label_by_id: dict[str, str],
    start_year: int | None,
    end_year: int | None,
    orcid_identifier_matches: dict[str, list[str]] | None = None,
    orcid_affiliation_matches: dict[str, dict[str, str]] | None = None,
    target_orcid: str = "",
) -> list[dict[str, object]]:
    if all_matches is None:
        all_matches = matches

    citation_by_pmid = {citation.pmid: citation for citation in citations}
    all_positions_by_pmid: dict[str, set[int]] = {}
    all_methods_by_pmid: dict[str, set[str]] = {}
    for match in all_matches:
        all_positions_by_pmid.setdefault(match.pmid, set()).add(match.position)
        all_methods_by_pmid.setdefault(match.pmid, set()).add(match.method)
    by_pmid: dict[str, dict[str, object]] = {}

    for match in matches:
        citation = citation_by_pmid.get(match.pmid)
        if citation is None:
            continue
        row = by_pmid.setdefault(
            match.pmid,
            {
                "pmid": citation.pmid,
                "pmid_link": f"https://pubmed.ncbi.nlm.nih.gov/{citation.pmid}/",
                "pmcid": citation.pmcid,
                "pmcid_link": f"https://pmc.ncbi.nlm.nih.gov/articles/{citation.pmcid}/" if citation.pmcid else None,
                "year": citation.print_year,
                "journal": citation.journal,
                "title": citation.title,
                "source_for_roles": citation.source_for_roles,
                "publication_types": ", ".join(citation.publication_types) if citation.publication_types else "(unknown)",
                "cluster_labels": set(),
                "affiliations": set(),
                "positions": set(),
                "notes": set(citation.notes),
            },
        )
        cluster_labels = cast(set[str], row["cluster_labels"])
        cluster_labels.add(cluster_label_by_id.get(match.cluster_id, match.cluster_id))
        affiliations = cast(set[str], row["affiliations"])
        affiliation = (match.author.affiliation or "").strip()
        if not affiliation:
            affiliation = "; ".join(_author_affiliation_blocks(match.author))
        if affiliation:
            affiliations.add(affiliation)
        positions = cast(set[int], row["positions"])
        positions.add(match.position)

    rendered: list[dict[str, object]] = []
    for row in by_pmid.values():
        pmid = str(row["pmid"])
        all_positions = all_positions_by_pmid.get(pmid, set())
        all_methods = all_methods_by_pmid.get(pmid, set())
        row_affiliations = cast(set[str], row["affiliations"])
        row_cluster_labels = cast(set[str], row["cluster_labels"])
        row_positions = cast(set[int], row["positions"])
        row_notes = cast(set[str], row["notes"])
        has_no_affiliation = not bool(row_affiliations)
        has_multiple_matches = len(all_positions) > 1
        has_initials_match = "initials" in all_methods
        review_reasons: list[str] = []
        if has_no_affiliation:
            review_reasons.append("No affiliation listed for matched author")
        if has_multiple_matches:
            review_reasons.append("Multiple matched author positions in citation")
        if has_initials_match:
            review_reasons.append("Matched by initials fallback")
        rendered.append(
            {
                "pmid": row["pmid"],
                "pmid_link": row["pmid_link"],
                "pmcid": row["pmcid"],
                "pmcid_link": row["pmcid_link"],
                "year": row["year"],
                "journal": row["journal"],
                "title": row["title"],
                "source_for_roles": row["source_for_roles"],
                "publication_types": row["publication_types"],
                "cluster_labels": " | ".join(sorted(row_cluster_labels)),
                "affiliations": " | ".join(sorted(row_affiliations)) if row_affiliations else "(no affiliation listed)",
                "authors_html": _to_author_list_html(citation_by_pmid[pmid], set(row_positions)),
                "source_badges": _citation_source_badges(citation_by_pmid[pmid]),
                "notes": "; ".join(sorted(row_notes)),
                "needs_review": bool(review_reasons),
                "review_reason": "; ".join(review_reasons),
                "in_window": bool(citation_in_window(citation_by_pmid[pmid], start_year, end_year)),
                **_orcid_row_fields(
                    pmid=pmid,
                    target_orcid=target_orcid,
                    orcid_identifier_matches=orcid_identifier_matches,
                    orcid_affiliation_matches=orcid_affiliation_matches,
                ),
            }
        )

    rendered.sort(key=_dict_row_sort_key)
    return rendered


def _build_cluster_citation_rows(
    *,
    citations: list[Citation],
    matches: list[AuthorMatch],
    orcid_identifier_matches: dict[str, list[str]] | None = None,
    orcid_affiliation_matches: dict[str, dict[str, str]] | None = None,
    target_orcid: str = "",
) -> dict[str, list[dict[str, object]]]:
    citation_by_pmid = {citation.pmid: citation for citation in citations}
    by_cluster: dict[str, dict[str, dict[str, object]]] = {}

    for match in matches:
        citation = citation_by_pmid.get(match.pmid)
        if citation is None:
            continue
        cluster_rows = by_cluster.setdefault(match.cluster_id, {})
        cluster_rows.setdefault(
            match.pmid,
            {
                "pmid": citation.pmid,
                "pmid_link": f"https://pubmed.ncbi.nlm.nih.gov/{citation.pmid}/",
                "year": citation.print_year,
                "journal": citation.journal,
                "title": citation.title,
                "source_badges": _citation_source_badges(citation),
                **_orcid_row_fields(
                    pmid=citation.pmid,
                    target_orcid=target_orcid,
                    orcid_identifier_matches=orcid_identifier_matches,
                    orcid_affiliation_matches=orcid_affiliation_matches,
                ),
            },
        )

    rendered: dict[str, list[dict[str, object]]] = {}
    for cluster_id, rows_by_pmid in by_cluster.items():
        rows = list(rows_by_pmid.values())
        rows.sort(key=_dict_row_sort_key)
        rendered[cluster_id] = rows
    return rendered


def _build_cluster_orcid_affiliation_matches(
    cluster_citation_rows: dict[str, list[dict[str, object]]],
) -> dict[str, list[dict[str, str]]]:
    by_cluster: dict[str, list[dict[str, str]]] = {}
    for cluster_id, rows in cluster_citation_rows.items():
        by_key: dict[str, dict[str, str]] = {}
        for row in rows:
            if not bool(row.get("orcid_affiliation_verified", False)):
                continue
            institution = str(row.get("orcid_affiliation_institution", "")).strip()
            level = str(row.get("orcid_affiliation_level", "")).strip().lower() or "likely"
            key = _institution_key(institution)
            if not institution or not key:
                continue
            existing = by_key.get(key)
            if existing is None or _orcid_affiliation_level_rank(level) > _orcid_affiliation_level_rank(existing["level"]):
                by_key[key] = {"key": key, "name": institution, "level": level}
        if by_key:
            by_cluster[cluster_id] = sorted(by_key.values(), key=lambda item: item["name"].lower())
    return by_cluster


def _build_cluster_affiliation_display_rows(
    *,
    clusters: list[Cluster],
    cluster_orcid_affiliation_matches: dict[str, list[dict[str, str]]],
    target_orcid: str,
) -> dict[str, list[dict[str, object]]]:
    by_cluster: dict[str, list[dict[str, object]]] = {}
    for cluster in clusters:
        matched_entries = cluster_orcid_affiliation_matches.get(cluster.cluster_id, [])
        rows: list[dict[str, object]] = []
        for affiliation in cluster.affiliations:
            key = _institution_key(affiliation)
            best_level = ""
            best_name = ""
            for entry in matched_entries:
                entry_key = str(entry.get("key", "")).strip()
                entry_level = str(entry.get("level", "")).strip().lower()
                entry_name = str(entry.get("name", "")).strip()
                if not key or not entry_key:
                    continue
                if key == entry_key:
                    candidate_level = entry_level
                elif (entry_key in key or key in entry_key) and _has_distinctive_institution_overlap(affiliation, entry_name):
                    candidate_level = "contains"
                else:
                    continue
                if _orcid_affiliation_level_rank(candidate_level) > _orcid_affiliation_level_rank(best_level):
                    best_level = candidate_level
                    best_name = entry_name

            matched = bool(best_level)
            rows.append(
                {
                    "name": affiliation,
                    "matched": matched,
                    "match_level": best_level,
                    "badge_class": _orcid_badge_class(
                        has_identifier_match=False,
                        affiliation_level=best_level,
                    )
                    if matched
                    else "orcid-badge",
                    "badge_title": _orcid_badge_title(
                        has_identifier_match=False,
                        identifier_label="",
                        affiliation_level=best_level,
                        affiliation_label=f"institution match ({best_name})" if best_name else "",
                    )
                    if matched
                    else "",
                    "orcid_record_url": f"https://orcid.org/{target_orcid}" if (target_orcid and matched) else None,
                }
            )
        by_cluster[cluster.cluster_id] = rows
    return by_cluster

def _sanitize_citation_affiliations(citations: list[Citation]) -> int:
    changed = 0
    for citation in citations:
        for author in citation.authors:
            original_affiliation = author.affiliation or ""
            original_blocks = _author_affiliation_blocks(author)
            cleaned_blocks: list[str] = []
            seen_blocks: set[str] = set()
            for block in original_blocks:
                cleaned = _normalize_affiliation_text(block)
                if not cleaned or cleaned in seen_blocks:
                    continue
                seen_blocks.add(cleaned)
                cleaned_blocks.append(cleaned)
            cleaned_affiliation = "; ".join(cleaned_blocks)
            if cleaned_blocks != original_blocks or cleaned_affiliation != original_affiliation:
                author.affiliation_blocks = cleaned_blocks
                author.affiliation = cleaned_affiliation
                changed += 1
    return changed


def _read_uploaded_text(upload: UploadFile | None) -> tuple[str, int]:
    if upload is None:
        return "", 0
    try:
        raw = upload.file.read(MAX_UPLOAD_BYTES + 1)
    except Exception:  # noqa: BLE001
        logger.exception("Failed to read uploaded PMID file: filename=%r", upload.filename)
        return "", 0
    if not raw:
        return "", 0
    if len(raw) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"Uploaded PMID file is too large. Maximum allowed size is {MAX_UPLOAD_BYTES} bytes."
        )
    return raw.decode("utf-8", errors="ignore"), len(raw)


def _uploaded_filename(upload: UploadFile | None) -> str:
    return (upload.filename or "").strip() if upload is not None else ""


def _record_usage_safe(**kwargs: Any) -> int | None:
    try:
        return record_usage_run(USAGE_DB_PATH, **kwargs)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Usage audit write skipped: %s", exc)
        return None


def _update_usage_safe(record_id: int | None, **kwargs: Any) -> None:
    if record_id is None:
        return
    try:
        update_usage_run(USAGE_DB_PATH, record_id, **kwargs)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Usage audit update skipped for record_id=%s: %s", record_id, exc)


def _admin_username() -> str:
    return os.getenv("ECHIDNA_ADMIN_USER", "").strip()


def _admin_password_hash() -> str:
    return os.getenv("ECHIDNA_ADMIN_PASSWORD_HASH", "").strip()


def _admin_password() -> str:
    return os.getenv("ECHIDNA_ADMIN_PASSWORD", "").strip()


def _require_admin_usage_access(credentials: HTTPBasicCredentials | None = Depends(admin_basic)) -> None:
    username = _admin_username()
    password_hash = _admin_password_hash()
    password = _admin_password()
    if not username or (not password_hash and not password):
        raise HTTPException(status_code=503, detail="Admin usage endpoint not configured")
    if credentials is None:
        raise HTTPException(
            status_code=401,
            detail="Unauthorized",
            headers={"WWW-Authenticate": "Basic"},
        )
    username_matches = secrets.compare_digest(credentials.username, username)
    if not username_matches:
        raise HTTPException(
            status_code=401,
            detail="Unauthorized",
            headers={"WWW-Authenticate": "Basic"},
        )
    try:
        if password_hash:
            password_matches = verify_password_against_hash(credentials.password, password_hash)
        else:
            password_matches = secrets.compare_digest(credentials.password, password)
    except ValueError as exc:
        logger.warning("Invalid admin password hash configuration: %s", exc)
        raise HTTPException(status_code=503, detail="Admin usage endpoint not configured") from exc
    if not password_matches:
        raise HTTPException(
            status_code=401,
            detail="Unauthorized",
            headers={"WWW-Authenticate": "Basic"},
        )


def _build_out_of_window_rows(
    *,
    citations: list[Citation],
    matches: list[AuthorMatch],
    start_year: int | None,
    end_year: int | None,
    forced_include_pmids: set[str] | None = None,
    orcid_identifier_matches: dict[str, list[str]] | None = None,
    orcid_affiliation_matches: dict[str, dict[str, str]] | None = None,
    target_orcid: str = "",
) -> list[dict[str, object]]:
    if start_year is None or end_year is None:
        return []
    if forced_include_pmids is None:
        forced_include_pmids = set()
    citation_by_pmid = {citation.pmid: citation for citation in citations}
    selected_pmids = {match.pmid for match in matches}
    rows: list[dict[str, object]] = []
    for pmid in selected_pmids:
        if pmid in forced_include_pmids:
            continue
        citation = citation_by_pmid.get(pmid)
        if citation is None:
            continue
        if citation_in_window(citation, start_year, end_year):
            continue
        rows.append(
            {
                "pmid": citation.pmid,
                "pmid_link": f"https://pubmed.ncbi.nlm.nih.gov/{citation.pmid}/",
                "pmcid": citation.pmcid,
                "pmcid_link": f"https://pmc.ncbi.nlm.nih.gov/articles/{citation.pmcid}/" if citation.pmcid else None,
                "year": citation.print_year,
                "journal": citation.journal,
                "title": citation.title,
                "source_badges": _citation_source_badges(citation),
                **_orcid_row_fields(
                    pmid=citation.pmid,
                    target_orcid=target_orcid,
                    orcid_identifier_matches=orcid_identifier_matches,
                    orcid_affiliation_matches=orcid_affiliation_matches,
                ),
            }
        )
    rows.sort(key=_dict_row_sort_key)
    return rows


def _build_excluded_citation_rows(
    *,
    citations: list[Citation],
    excluded_pmids: set[str],
    excluded_reasons: dict[str, str],
    matches: list[AuthorMatch],
    orcid_identifier_matches: dict[str, list[str]] | None = None,
    orcid_affiliation_matches: dict[str, dict[str, str]] | None = None,
    target_orcid: str = "",
) -> list[dict[str, object]]:
    citation_by_pmid = {citation.pmid: citation for citation in citations}
    matched_pmids = {match.pmid for match in matches}
    rows: list[dict[str, object]] = []
    for pmid in sorted(excluded_pmids):
        if pmid not in matched_pmids:
            continue
        citation = citation_by_pmid.get(pmid)
        if citation is None:
            continue
        rows.append(
            {
                "pmid": citation.pmid,
                "pmid_link": f"https://pubmed.ncbi.nlm.nih.gov/{citation.pmid}/",
                "pmcid": citation.pmcid,
                "pmcid_link": f"https://pmc.ncbi.nlm.nih.gov/articles/{citation.pmcid}/" if citation.pmcid else None,
                "year": citation.print_year,
                "journal": citation.journal,
                "title": citation.title,
                "reason": excluded_reasons.get(citation.pmid, "Excluded by pre-disambiguation filter"),
                "source_badges": _citation_source_badges(citation),
                **_orcid_row_fields(
                    pmid=citation.pmid,
                    target_orcid=target_orcid,
                    orcid_identifier_matches=orcid_identifier_matches,
                    orcid_affiliation_matches=orcid_affiliation_matches,
                ),
            }
        )
    rows.sort(key=_dict_row_sort_key)
    return rows


def _partition_excluded_citation_rows(
    rows: list[dict[str, object]],
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    by_publication_type: list[dict[str, object]] = []
    by_date_window: list[dict[str, object]] = []
    other: list[dict[str, object]] = []
    for row in rows:
        reason = str(row.get("reason", "")).strip().lower()
        if reason.startswith("excluded by publication type"):
            by_publication_type.append(row)
            continue
        if reason.startswith("outside year window"):
            by_date_window.append(row)
            continue
        other.append(row)
    return by_publication_type, by_date_window, other


def _filter_matches_for_analysis(
    *,
    run: RunState,
    selected_cluster_ids: set[str],
    selected_pmids: set[str],
    pmid_filter_enabled: int,
    forced_excluded_pmids: set[str],
) -> list[AuthorMatch]:
    return [
        match
        for match in run.matches
        if (
            (match.cluster_id in selected_cluster_ids or match.pmid in forced_excluded_pmids)
            and (pmid_filter_enabled == 0 or match.pmid in selected_pmids)
            and (match.pmid not in run.excluded_pmids or match.pmid in forced_excluded_pmids)
        )
    ]


def _save_run(run_id: str, run: RunState) -> None:
    _prune_expired_runs()
    RUNS[run_id] = run
    try:
        _save_run_to_disk(run_id, run)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Run-state disk persistence skipped for run_id=%s: %s", run_id, exc)


def _run_file(run_id: str) -> Path:
    return RUNS_DIR / f"{run_id}.json"


def _run_is_expired(run: RunState, *, now: float | None = None) -> bool:
    if now is None:
        now = time.time()
    return run.created_at <= 0 or (now - run.created_at) > RUN_TTL_SECONDS


def _delete_run_file(run_id: str) -> None:
    if not RUN_ID_PATTERN.fullmatch(run_id):
        return
    try:
        _run_file(run_id).unlink(missing_ok=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Run-state disk cleanup skipped for run_id=%s: %s", run_id, exc)


def _prune_expired_runs(*, now: float | None = None) -> None:
    if now is None:
        now = time.time()

    expired_run_ids = [run_id for run_id, run in RUNS.items() if _run_is_expired(run, now=now)]
    for run_id in expired_run_ids:
        RUNS.pop(run_id, None)
        _delete_run_file(run_id)

    if not RUNS_DIR.exists():
        return

    ttl_cutoff = now - RUN_TTL_SECONDS
    for target in RUNS_DIR.glob("*.json"):
        if target.stat().st_mtime > ttl_cutoff:
            continue
        run_id = target.stem
        if not RUN_ID_PATTERN.fullmatch(run_id):
            continue
        try:
            target.unlink()
        except Exception as exc:  # noqa: BLE001
            logger.warning("Run-state file cleanup skipped for run_id=%s: %s", run_id, exc)


def _save_run_to_disk(run_id: str, run: RunState) -> None:
    if not RUN_ID_PATTERN.fullmatch(run_id):
        logger.warning("Refused to persist run with invalid id format: %r", run_id)
        return
    RUNS_DIR.mkdir(parents=True, exist_ok=True)
    os.chmod(RUNS_DIR, 0o700)
    target = _run_file(run_id)
    temp = target.with_suffix(".json.tmp")
    payload = _serialize_run(run)
    temp.write_text(json.dumps(payload, separators=(",", ":")), encoding="utf-8")
    os.chmod(temp, 0o600)
    temp.replace(target)
    os.chmod(target, 0o600)


def _load_run_from_disk(run_id: str) -> RunState | None:
    if not RUN_ID_PATTERN.fullmatch(run_id):
        return None
    try:
        target = _run_file(run_id)
        if not target.exists():
            return None
        payload = json.loads(target.read_text(encoding="utf-8"))
    except Exception as exc:  # noqa: BLE001
        logger.warning("Run-state disk load skipped for run_id=%s: %s", run_id, exc)
        return None
    try:
        created_at_fallback = target.stat().st_mtime
        run = _deserialize_run(payload, created_at_fallback=created_at_fallback)
    except Exception:  # noqa: BLE001
        logger.exception("Failed to parse run-state payload for run_id=%s", run_id)
        return None
    if _run_is_expired(run):
        _delete_run_file(run_id)
        return None
    return run


def _serialize_run(run: RunState) -> dict[str, Any]:
    return {
        "created_at": run.created_at,
        "author_name": run.author_name,
        "target_orcid": run.target_orcid,
        "start_year": run.start_year,
        "end_year": run.end_year,
        "excluded_type_terms": list(run.excluded_type_terms),
        "citations": [_serialize_citation(citation) for citation in run.citations],
        "clusters": [_serialize_cluster(cluster) for cluster in run.clusters],
        "matches": [_serialize_match(match) for match in run.matches],
        "excluded_pmids": sorted(run.excluded_pmids),
        "excluded_reasons": dict(run.excluded_reasons),
        "orcid_identifier_matches": {k: list(v) for k, v in run.orcid_identifier_matches.items()},
        "orcid_affiliation_matches": {k: {str(sub_k): str(sub_v) for sub_k, sub_v in value.items()} for k, value in run.orcid_affiliation_matches.items()},
        "orcid_sync_error": run.orcid_sync_error,
    }


def _deserialize_run(payload: dict[str, Any], *, created_at_fallback: float | None = None) -> RunState:
    start_raw = payload.get("start_year")
    end_raw = payload.get("end_year")
    created_at_raw = payload.get("created_at", created_at_fallback)
    try:
        created_at = float(created_at_raw) if created_at_raw is not None else time.time()
    except (TypeError, ValueError):
        created_at = time.time()
    return RunState(
        author_name=str(payload["author_name"]),
        target_orcid=str(payload.get("target_orcid", "")),
        start_year=(int(start_raw) if start_raw is not None else None),
        end_year=(int(end_raw) if end_raw is not None else None),
        excluded_type_terms=[str(v) for v in payload.get("excluded_type_terms", [])],
        citations=[_deserialize_citation(item) for item in payload.get("citations", [])],
        clusters=[_deserialize_cluster(item) for item in payload.get("clusters", [])],
        matches=[_deserialize_match(item) for item in payload.get("matches", [])],
        excluded_pmids={str(v) for v in payload.get("excluded_pmids", [])},
        excluded_reasons={str(k): str(v) for k, v in dict(payload.get("excluded_reasons", {})).items()},
        orcid_identifier_matches={
            str(k): [str(item) for item in values]
            for k, values in dict(payload.get("orcid_identifier_matches", {})).items()
            if isinstance(values, list)
        },
        orcid_affiliation_matches={
            str(k): {str(sub_k): str(sub_v) for sub_k, sub_v in dict(value).items()}
            for k, value in dict(payload.get("orcid_affiliation_matches", {})).items()
            if isinstance(value, dict)
        },
        orcid_sync_error=(str(payload.get("orcid_sync_error", "")).strip() or None),
        created_at=created_at,
    )


def _split_affiliation_blocks(raw_affiliation: str) -> list[str]:
    return [piece.strip() for piece in re.split(r"[;|]+", raw_affiliation) if piece.strip()]


def _author_affiliation_blocks(author: Any) -> list[str]:
    raw_blocks = getattr(author, "affiliation_blocks", [])
    blocks = [str(value).strip() for value in raw_blocks if str(value).strip()] if isinstance(raw_blocks, list) else []
    if blocks:
        return blocks
    return _split_affiliation_blocks(str(getattr(author, "affiliation", "")))


def _serialize_author(author: Any) -> dict[str, Any]:
    affiliation_blocks = _author_affiliation_blocks(author)
    return {
        "position": int(author.position),
        "last_name": str(author.last_name),
        "fore_name": str(author.fore_name),
        "initials": str(author.initials),
        "affiliation": str(author.affiliation),
        "affiliation_blocks": affiliation_blocks,
        "orcid": str(getattr(author, "orcid", "")),
    }


def _deserialize_author(payload: dict[str, Any]) -> Any:
    from app.models import Author

    raw_blocks = payload.get("affiliation_blocks", [])
    if isinstance(raw_blocks, list):
        affiliation_blocks = [str(value).strip() for value in raw_blocks if str(value).strip()]
    else:
        affiliation_blocks = []
    if not affiliation_blocks:
        affiliation_blocks = _split_affiliation_blocks(str(payload.get("affiliation", "")))

    return Author(
        position=int(payload["position"]),
        last_name=str(payload.get("last_name", "")),
        fore_name=str(payload.get("fore_name", "")),
        initials=str(payload.get("initials", "")),
        affiliation=str(payload.get("affiliation", "")).strip() or "; ".join(affiliation_blocks),
        affiliation_blocks=affiliation_blocks,
        orcid=str(payload.get("orcid", "")),
    )


def _serialize_citation(citation: Citation) -> dict[str, Any]:
    return {
        "pmid": citation.pmid,
        "pmcid": citation.pmcid,
        "title": citation.title,
        "journal": citation.journal,
        "print_year": citation.print_year,
        "print_date": citation.print_date,
        "authors": [_serialize_author(author) for author in citation.authors],
        "publication_types": list(citation.publication_types),
        "source_for_roles": citation.source_for_roles,
        "co_first_positions": sorted(citation.co_first_positions),
        "co_senior_positions": sorted(citation.co_senior_positions),
        "notes": list(citation.notes),
        "doi": citation.doi,
        "source_tags": sorted(citation.source_tags),
        "update_in_pmids": sorted(citation.update_in_pmids),
        "update_of_pmids": sorted(citation.update_of_pmids),
    }


def _deserialize_citation(payload: dict[str, Any]) -> Citation:
    return Citation(
        pmid=str(payload.get("pmid", "")),
        pmcid=payload.get("pmcid"),
        title=str(payload.get("title", "")),
        journal=str(payload.get("journal", "")),
        print_year=payload.get("print_year"),
        print_date=str(payload.get("print_date", "")),
        authors=[_deserialize_author(item) for item in payload.get("authors", [])],
        publication_types=[str(v) for v in payload.get("publication_types", [])],
        source_for_roles=str(payload.get("source_for_roles", "pubmed")),  # type: ignore[arg-type]
        co_first_positions={int(v) for v in payload.get("co_first_positions", [])},
        co_senior_positions={int(v) for v in payload.get("co_senior_positions", [])},
        notes=[str(v) for v in payload.get("notes", [])],
        doi=(str(payload.get("doi", "")).strip() or None),
        source_tags={str(v) for v in payload.get("source_tags", [])},
        update_in_pmids={str(v) for v in payload.get("update_in_pmids", [])},
        update_of_pmids={str(v) for v in payload.get("update_of_pmids", [])},
    )


def _serialize_cluster(cluster: Cluster) -> dict[str, Any]:
    return {
        "cluster_id": cluster.cluster_id,
        "label": cluster.label,
        "mention_count": cluster.mention_count,
        "years": list(cluster.years),
        "affiliations": list(cluster.affiliations),
    }


def _deserialize_cluster(payload: dict[str, Any]) -> Cluster:
    return Cluster(
        cluster_id=str(payload.get("cluster_id", "")),
        label=str(payload.get("label", "")),
        mention_count=int(payload.get("mention_count", 0)),
        years=[int(v) for v in payload.get("years", [])],
        affiliations=[str(v) for v in payload.get("affiliations", [])],
    )


def _serialize_match(match: AuthorMatch) -> dict[str, Any]:
    return {
        "pmid": match.pmid,
        "position": match.position,
        "method": match.method,
        "author": _serialize_author(match.author),
        "cluster_id": match.cluster_id,
    }


def _deserialize_match(payload: dict[str, Any]) -> AuthorMatch:
    return AuthorMatch(
        pmid=str(payload.get("pmid", "")),
        position=int(payload.get("position", 0)),
        method=str(payload.get("method", "initials")),  # type: ignore[arg-type]
        author=_deserialize_author(payload.get("author", {})),
        cluster_id=str(payload.get("cluster_id", "")),
    )


def _load_run(run_id: str) -> RunState:
    _prune_expired_runs()
    run = RUNS.get(run_id)
    if run is not None:
        if _run_is_expired(run):
            RUNS.pop(run_id, None)
            _delete_run_file(run_id)
        else:
            return run
    run = _load_run_from_disk(run_id)
    if run is None:
        raise HTTPException(status_code=404, detail="Run not found or expired")
    RUNS[run_id] = run
    return run


def _index_context(
    *,
    author_name: str,
    author_orcid: str,
    start_year: int,
    end_year: int,
    pmids: str,
    all_years: bool,
    include_preprints: bool,
    extra_excluded_types: str,
    default_excluded_type_options: list[str],
    selected_default_excluded_types: list[str],
    error: str | None,
    notice: str | None = None,
) -> dict[str, object]:
    return {
        "author_name": author_name,
        "author_orcid": author_orcid,
        "start_year": start_year,
        "end_year": end_year,
        "pmids": pmids,
        "all_years": all_years,
        "include_preprints": include_preprints,
        "extra_excluded_types": extra_excluded_types,
        "default_excluded_type_options": default_excluded_type_options,
        "selected_default_excluded_types": selected_default_excluded_types,
        "year_options": _year_options(),
        "error": error,
        "notice": notice,
    }


def _usage_record_replay_context(record: dict[str, Any]) -> dict[str, object]:
    default_excluded_type_options = _publication_type_exclusion_options()
    all_years = bool(record["all_years"])
    start_year = record["start_year"]
    end_year = record["end_year"]
    if not isinstance(start_year, int):
        start_year = date.today().year - 5
    if not isinstance(end_year, int):
        end_year = date.today().year

    excluded_type_terms = _dedupe_terms([str(item) for item in record["excluded_type_terms"]])
    selected_default_excluded_types = [term for term in default_excluded_type_options if term in excluded_type_terms]
    extra_excluded_types = "\n".join(
        term
        for term in excluded_type_terms
        if term not in default_excluded_type_options and term != PREPRINT_PUBLICATION_TYPE
    )

    uploaded_filename = str(record["uploaded_filename"]).strip()
    notice = f"Loaded diagnostic record {record['id']} for replay."
    if uploaded_filename:
        notice += f" Original upload {uploaded_filename!r} has been restored into the PMID text box."

    return _index_context(
        author_name=str(record["author_name"]),
        author_orcid=str(record["target_orcid"]),
        start_year=start_year,
        end_year=end_year,
        pmids=str(record["submitted_pmids_text"]),
        all_years=all_years,
        include_preprints=bool(record["include_preprints"]),
        extra_excluded_types=extra_excluded_types,
        default_excluded_type_options=default_excluded_type_options,
        selected_default_excluded_types=selected_default_excluded_types,
        error=None,
        notice=notice,
    )


def _acquire_disambiguation_slot() -> bool:
    return DISAMBIGUATION_SEMAPHORE.acquire(blocking=False)


def _busy_disambiguation_response(
    request: Request,
    *,
    author_name: str,
    author_orcid: str,
    start_year: int,
    end_year: int,
    pmids: str,
    all_years: bool,
    include_preprints: bool,
    extra_excluded_types: str,
    default_excluded_type_options: list[str],
    selected_default_excluded_types: list[str],
) -> HTMLResponse:
    return templates.TemplateResponse(
        request,
        "index.html",
        _index_context(
            author_name=author_name,
            author_orcid=author_orcid,
            start_year=start_year,
            end_year=end_year,
            pmids=pmids,
            all_years=all_years,
            include_preprints=include_preprints,
            extra_excluded_types=extra_excluded_types,
            default_excluded_type_options=default_excluded_type_options,
            selected_default_excluded_types=selected_default_excluded_types,
            error="Server busy, retry shortly.",
        ),
        status_code=503,
        headers={"Retry-After": str(BUSY_RETRY_AFTER_SECONDS)},
    )


def _render_analysis_result(
    request: Request,
    *,
    run: RunState,
    run_id: str,
    selected_cluster_ids: set[str],
    selected_pmids: set[str],
    selected_excluded_pmids: set[str],
    pmid_filter_enabled: int,
    analysis: AnalysisResult,
    start_year: int | None,
    end_year: int | None,
    force_report: bool = False,
    out_of_window_rows: list[dict[str, object]] | None = None,
) -> HTMLResponse:
    if analysis.uncertain_indices and not force_report:
        logger.info(
            "Analysis requires manual review for %d citation(s)",
            len(analysis.uncertain_indices),
        )
        row_data = _with_row_render_fields(
            analysis,
            orcid_identifier_matches=run.orcid_identifier_matches,
            orcid_affiliation_matches=run.orcid_affiliation_matches,
            target_orcid=run.target_orcid,
            orcid_identifier_checked=_orcid_identifier_crosscheck_available(
                target_orcid=run.target_orcid,
                orcid_sync_error=run.orcid_sync_error,
            ),
        )
        uncertain_rows = [row_data[i] for i in analysis.uncertain_indices]
        return templates.TemplateResponse(
            request,
            "review_uncertain.html",
            {
                "run_id": run_id,
                "selected_cluster_ids": sorted(selected_cluster_ids),
                "selected_pmids": sorted(selected_pmids),
                "selected_excluded_pmids": sorted(selected_excluded_pmids),
                "pmid_filter_enabled": pmid_filter_enabled,
                "rows": uncertain_rows,
            },
        )

    summary = format_summary(analysis.rows, start_year, end_year)
    row_data = _with_row_render_fields(
        analysis,
        orcid_identifier_matches=run.orcid_identifier_matches,
        orcid_affiliation_matches=run.orcid_affiliation_matches,
        target_orcid=run.target_orcid,
        orcid_identifier_checked=_orcid_identifier_crosscheck_available(
            target_orcid=run.target_orcid,
            orcid_sync_error=run.orcid_sync_error,
        ),
    )
    accepted_uncertain_pmids = sorted(
        {
            row.citation.pmid
            for row in analysis.rows
            if row.uncertainty_reasons and row.include
        }
    )
    # Final report grouping:
    # 1) excluded rows
    # 2) included rows where PMC metadata was unavailable (PubMed fallback)
    # 3) included rows where PMC author contribution metadata was available
    excluded_rows = [row for row in row_data if not row["include"]]
    included_pubmed_rows = [row for row in row_data if row["include"] and row["role_source"] != "pmc"]
    included_pmc_rows = [row for row in row_data if row["include"] and row["role_source"] == "pmc"]
    report_rows = excluded_rows + included_pubmed_rows + included_pmc_rows
    return templates.TemplateResponse(
        request,
        "report.html",
        {
            "summary": summary,
            "rows": report_rows,
            "out_of_window_rows": out_of_window_rows or [],
            "start_year": start_year,
            "end_year": end_year,
            "year_window_label": _year_window_label(start_year, end_year),
            "run_id": run_id,
            "selected_cluster_ids": sorted(selected_cluster_ids),
            "selected_pmids": sorted(selected_pmids),
            "selected_excluded_pmids": sorted(selected_excluded_pmids),
            "pmid_filter_enabled": pmid_filter_enabled,
            "accepted_uncertain_pmids": accepted_uncertain_pmids,
        },
    )


@app.get("/", response_class=HTMLResponse)
def index(request: Request) -> HTMLResponse:
    year = date.today().year
    default_excluded_type_options = _publication_type_exclusion_options()
    return templates.TemplateResponse(
        request,
        "index.html",
        {
            "author_name": "",
            "author_orcid": "",
            "start_year": year - 5,
            "end_year": year,
            "pmids": "",
            "all_years": False,
            "include_preprints": False,
            "extra_excluded_types": "",
            "default_excluded_type_options": default_excluded_type_options,
            "selected_default_excluded_types": list(default_excluded_type_options),
            "year_options": _year_options(),
            "error": None,
            "notice": None,
        },
    )


@app.get("/orcid-search")
def orcid_search(
    author_name: str = Query(..., min_length=2),
    limit: int = Query(default=25, ge=1, le=50),
) -> dict[str, object]:
    query = author_name.strip()
    if len(query) < 2:
        raise HTTPException(status_code=400, detail="Enter at least 2 characters for ORCiD search")

    try:
        matches = orcid_client.search_profiles(query, limit=limit)
    except Exception as exc:  # noqa: BLE001
        logger.warning("ORCiD profile search failed for %r: %s", query, exc)
        raise HTTPException(status_code=502, detail="ORCiD search unavailable") from exc

    return {
        "author_name": query,
        "matches": matches,
    }


@app.get("/admin/usage")
def admin_usage(
    _: None = Depends(_require_admin_usage_access),
    limit: int = Query(default=50, ge=1, le=200),
    author_name: str = Query(default=""),
    status: str = Query(default=""),
) -> dict[str, object]:
    try:
        items = fetch_usage_runs(
            USAGE_DB_PATH,
            limit=limit,
            author_name=author_name,
            status=status,
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning("Usage audit fetch failed: %s", exc)
        raise HTTPException(status_code=500, detail="Usage audit unavailable") from exc
    for item in items:
        item["rerun_url"] = f"/admin/usage/{item['id']}/rerun"
    return {"items": items}


@app.get("/admin/usage/{record_id}/rerun", response_class=HTMLResponse)
def admin_usage_rerun(
    request: Request,
    record_id: int,
    _: None = Depends(_require_admin_usage_access),
) -> HTMLResponse:
    try:
        record = fetch_usage_run_by_id(USAGE_DB_PATH, record_id)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Usage audit replay lookup failed for record_id=%s: %s", record_id, exc)
        raise HTTPException(status_code=500, detail="Usage audit unavailable") from exc
    if record is None:
        raise HTTPException(status_code=404, detail="Usage record not found")
    return templates.TemplateResponse(
        request,
        "index.html",
        _usage_record_replay_context(record),
    )


@app.get("/echidna.jpg", include_in_schema=False)
def landing_image() -> FileResponse:
    if not LANDING_IMAGE_PATH.is_file():
        raise HTTPException(status_code=404, detail="echidna.jpg not found")
    return FileResponse(LANDING_IMAGE_PATH, media_type="image/jpeg")


@app.get("/favicon.ico", include_in_schema=False)
def favicon() -> FileResponse:
    if not FAVICON_PATH.is_file():
        raise HTTPException(status_code=404, detail="favicon.ico not found")
    return FileResponse(FAVICON_PATH, media_type="image/x-icon")


@app.post("/disambiguate", response_class=HTMLResponse)
def disambiguate(
    request: Request,
    author_name: str = Form(...),
    author_orcid: str = Form(default=""),
    pmids: str = Form(default=""),
    pmid_file: UploadFile | None = File(default=None),
    start_year: int = Form(...),
    end_year: int = Form(...),
    all_years: str | None = Form(default=None),
    include_preprints: str | None = Form(default=None),
    default_excluded_types: list[str] = Form(default=[]),
    default_excluded_types_present: str | None = Form(default=None),
    extra_excluded_types: str = Form(default=""),
) -> HTMLResponse:
    all_years_enabled = all_years is not None
    include_preprints_enabled = include_preprints is not None
    uploaded_filename = _uploaded_filename(pmid_file)
    default_excluded_type_options = _publication_type_exclusion_options()
    if default_excluded_types_present is None:
        # Backward compatibility for clients that do not submit the checklist fields.
        selected_default_excluded_types = list(default_excluded_type_options)
    else:
        selected_default_excluded_types = _selected_default_excluded_type_terms(
            default_excluded_types, default_excluded_type_options
        )
    excluded_type_terms = _dedupe_terms(selected_default_excluded_types + parse_type_terms(extra_excluded_types))
    if not include_preprints_enabled:
        excluded_type_terms = _dedupe_terms([PREPRINT_PUBLICATION_TYPE] + excluded_type_terms)
    logger.info(
        "Disambiguation request received: author=%r, target_orcid=%r, start_year=%s, end_year=%s, all_years=%s, include_preprints=%s, excluded_type_terms=%d",
        author_name,
        author_orcid,
        start_year,
        end_year,
        all_years_enabled,
        include_preprints_enabled,
        len(excluded_type_terms),
    )
    try:
        target = parse_target_name(author_name)
        target_orcid = parse_target_orcid(author_orcid)
    except ValueError as exc:
        _record_usage_safe(
            run_id="",
            author_name=author_name,
            target_orcid=author_orcid,
            start_year=start_year,
            end_year=end_year,
            all_years=all_years_enabled,
            include_preprints=include_preprints_enabled,
            excluded_type_terms=excluded_type_terms,
            submitted_pmids_text=pmids,
            uploaded_filename=uploaded_filename,
            uploaded_size_bytes=0,
            parsed_pmids=[],
            status="rejected_invalid_author",
            error_message=str(exc),
        )
        return templates.TemplateResponse(
            request,
            "index.html",
            _index_context(
                author_name=author_name,
                author_orcid=author_orcid,
                start_year=start_year,
                end_year=end_year,
                pmids=pmids,
                all_years=all_years_enabled,
                include_preprints=include_preprints_enabled,
                extra_excluded_types=extra_excluded_types,
                default_excluded_type_options=default_excluded_type_options,
                selected_default_excluded_types=selected_default_excluded_types,
                error=str(exc),
            ),
            status_code=400,
        )

    try:
        pmid_file_text, uploaded_size_bytes = _read_uploaded_text(pmid_file)
    except ValueError as exc:
        _record_usage_safe(
            run_id="",
            author_name=author_name,
            target_orcid=target_orcid,
            start_year=start_year,
            end_year=end_year,
            all_years=all_years_enabled,
            include_preprints=include_preprints_enabled,
            excluded_type_terms=excluded_type_terms,
            submitted_pmids_text=pmids,
            uploaded_filename=uploaded_filename,
            uploaded_size_bytes=MAX_UPLOAD_BYTES + 1,
            parsed_pmids=[],
            status="rejected_upload_too_large",
            error_message=str(exc),
        )
        return templates.TemplateResponse(
            request,
            "index.html",
            _index_context(
                author_name=author_name,
                author_orcid=author_orcid,
                start_year=start_year,
                end_year=end_year,
                pmids=pmids,
                all_years=all_years_enabled,
                include_preprints=include_preprints_enabled,
                extra_excluded_types=extra_excluded_types,
                default_excluded_type_options=default_excluded_type_options,
                selected_default_excluded_types=selected_default_excluded_types,
                error=str(exc),
            ),
            status_code=413,
        )

    combined_pmids_text = "\n".join(part for part in (pmids, pmid_file_text) if part.strip())
    parsed_pmids = parse_pmids(combined_pmids_text)
    logger.info("Parsed %d PMID(s) from input", len(parsed_pmids))
    if not parsed_pmids:
        _record_usage_safe(
            run_id="",
            author_name=author_name,
            target_orcid=target_orcid,
            start_year=start_year,
            end_year=end_year,
            all_years=all_years_enabled,
            include_preprints=include_preprints_enabled,
            excluded_type_terms=excluded_type_terms,
            submitted_pmids_text=combined_pmids_text,
            uploaded_filename=uploaded_filename,
            uploaded_size_bytes=uploaded_size_bytes,
            parsed_pmids=[],
            status="rejected_no_pmids",
            error_message="No PMIDs found in text input or uploaded file",
        )
        return templates.TemplateResponse(
            request,
            "index.html",
            _index_context(
                author_name=author_name,
                author_orcid=author_orcid,
                start_year=start_year,
                end_year=end_year,
                pmids=pmids,
                all_years=all_years_enabled,
                include_preprints=include_preprints_enabled,
                extra_excluded_types=extra_excluded_types,
                default_excluded_type_options=default_excluded_type_options,
                selected_default_excluded_types=selected_default_excluded_types,
                error="No PMIDs found in text input or uploaded file",
            ),
            status_code=400,
        )

    if len(parsed_pmids) > MAX_PMID_COUNT:
        limit_error = (
            f"Too many PMIDs submitted ({len(parsed_pmids)}). "
            f"Maximum allowed per run is {MAX_PMID_COUNT}. Please split the input into smaller batches."
        )
        _record_usage_safe(
            run_id="",
            author_name=author_name,
            target_orcid=target_orcid,
            start_year=start_year,
            end_year=end_year,
            all_years=all_years_enabled,
            include_preprints=include_preprints_enabled,
            excluded_type_terms=excluded_type_terms,
            submitted_pmids_text=combined_pmids_text,
            uploaded_filename=uploaded_filename,
            uploaded_size_bytes=uploaded_size_bytes,
            parsed_pmids=parsed_pmids,
            status="rejected_pmid_limit",
            error_message=limit_error,
        )
        return templates.TemplateResponse(
            request,
            "index.html",
            _index_context(
                author_name=author_name,
                author_orcid=author_orcid,
                start_year=start_year,
                end_year=end_year,
                pmids=pmids,
                all_years=all_years_enabled,
                include_preprints=include_preprints_enabled,
                extra_excluded_types=extra_excluded_types,
                default_excluded_type_options=default_excluded_type_options,
                selected_default_excluded_types=selected_default_excluded_types,
                error=limit_error,
            ),
            status_code=400,
        )

    if not all_years_enabled and start_year > end_year:
        _record_usage_safe(
            run_id="",
            author_name=author_name,
            target_orcid=target_orcid,
            start_year=start_year,
            end_year=end_year,
            all_years=all_years_enabled,
            include_preprints=include_preprints_enabled,
            excluded_type_terms=excluded_type_terms,
            submitted_pmids_text=combined_pmids_text,
            uploaded_filename=uploaded_filename,
            uploaded_size_bytes=uploaded_size_bytes,
            parsed_pmids=parsed_pmids,
            status="rejected_year_window",
            error_message="Start Year must be less than or equal to End Year",
        )
        return templates.TemplateResponse(
            request,
            "index.html",
            _index_context(
                author_name=author_name,
                author_orcid=author_orcid,
                start_year=start_year,
                end_year=end_year,
                pmids=pmids,
                all_years=all_years_enabled,
                include_preprints=include_preprints_enabled,
                extra_excluded_types=extra_excluded_types,
                default_excluded_type_options=default_excluded_type_options,
                selected_default_excluded_types=selected_default_excluded_types,
                error="Start Year must be less than or equal to End Year",
            ),
            status_code=400,
        )

    if not _acquire_disambiguation_slot():
        _record_usage_safe(
            run_id="",
            author_name=author_name,
            target_orcid=target_orcid,
            start_year=start_year,
            end_year=end_year,
            all_years=all_years_enabled,
            include_preprints=include_preprints_enabled,
            excluded_type_terms=excluded_type_terms,
            submitted_pmids_text=combined_pmids_text,
            uploaded_filename=uploaded_filename,
            uploaded_size_bytes=uploaded_size_bytes,
            parsed_pmids=parsed_pmids,
            status="rejected_busy",
            error_message="Server busy, retry shortly.",
        )
        return _busy_disambiguation_response(
            request,
            author_name=author_name,
            author_orcid=author_orcid,
            start_year=start_year,
            end_year=end_year,
            pmids=pmids,
            all_years=all_years_enabled,
            include_preprints=include_preprints_enabled,
            extra_excluded_types=extra_excluded_types,
            default_excluded_type_options=default_excluded_type_options,
            selected_default_excluded_types=selected_default_excluded_types,
        )

    usage_record_id = _record_usage_safe(
        run_id="",
        author_name=author_name,
        target_orcid=target_orcid,
        start_year=start_year,
        end_year=end_year,
        all_years=all_years_enabled,
        include_preprints=include_preprints_enabled,
        excluded_type_terms=excluded_type_terms,
        submitted_pmids_text=combined_pmids_text,
        uploaded_filename=uploaded_filename,
        uploaded_size_bytes=uploaded_size_bytes,
        parsed_pmids=parsed_pmids,
        status="accepted",
    )

    try:
        citations: list[Citation] = []
        errors: list[str] = []
        for chunk in split_chunks(parsed_pmids):
            try:
                citations.extend(client.fetch_citations(chunk))
            except Exception as exc:  # noqa: BLE001
                errors.append(str(exc))
                logger.exception("Chunk fetch failed for %d PMID(s)", len(chunk))
        user_provided_pmids = set(parsed_pmids)
        litsense_pmid_pmids: set[str] = set()
        litsense_orcid_pmids: set[str] = set()
        if parsed_pmids:
            seed_pmid = parsed_pmids[0]
            try:
                litsense_pmid_pmids = client.fetch_litsense_pmids_by_query(seed_pmid=seed_pmid, author_name=author_name)
                logger.info(
                    "LitSense seed query completed: seed=%s, returned=%d PMID(s)",
                    seed_pmid,
                    len(litsense_pmid_pmids),
                )
            except Exception as exc:  # noqa: BLE001
                logger.warning("LitSense seed query failed for seed=%s author=%r: %s", seed_pmid, author_name, exc)
                errors.append("LitSense PMID query unavailable")
        if target_orcid:
            try:
                litsense_orcid_pmids = client.fetch_litsense_pmids_by_orcid(target_orcid=target_orcid)
                logger.info(
                    "LitSense ORCiD query completed: orcid=%s, returned=%d PMID(s)",
                    target_orcid,
                    len(litsense_orcid_pmids),
                )
            except Exception as exc:  # noqa: BLE001
                logger.warning("LitSense ORCiD query failed for %s: %s", target_orcid, exc)
                errors.append("LitSense ORCiD query unavailable")

        existing_pmids = {citation.pmid for citation in citations}
        litsense_extra_pmids = sorted(
            (litsense_pmid_pmids | litsense_orcid_pmids) - user_provided_pmids - existing_pmids,
            key=int,
        )
        if litsense_extra_pmids:
            logger.info("Fetching %d LitSense-only PMID(s) for verification flow", len(litsense_extra_pmids))
            for chunk in split_chunks(litsense_extra_pmids):
                try:
                    citations.extend(client.fetch_citations(chunk))
                except Exception as exc:  # noqa: BLE001
                    errors.append(str(exc))
                    logger.exception("LitSense-only chunk fetch failed for %d PMID(s)", len(chunk))

        cleaned_affiliations = _sanitize_citation_affiliations(citations)
        if cleaned_affiliations:
            logger.info("Post-fetch affiliation sanitation updated %d author affiliation value(s)", cleaned_affiliations)

        for citation in citations:
            tags = set(citation.source_tags)
            if citation.pmid in user_provided_pmids:
                tags.add("user")
            if citation.pmid in litsense_pmid_pmids or citation.pmid in litsense_orcid_pmids:
                tags.add("litsense")
            citation.source_tags = tags

        orcid_identifier_matches: dict[str, list[str]] = {}
        orcid_affiliation_matches: dict[str, dict[str, str]] = {}
        orcid_sync_warnings: list[str] = []
        if target_orcid:
            try:
                orcid_identifier_matches = orcid_client.match_citations(target_orcid, citations)
                logger.info(
                    "ORCiD identifier cross-check completed: matched %d/%d citation(s)",
                    len(orcid_identifier_matches),
                    len(citations),
                )
            except Exception as exc:  # noqa: BLE001
                logger.warning("ORCiD identifier cross-check failed for %s: %s", target_orcid, exc)
                orcid_sync_warnings.append("identifier cross-check unavailable")

        orcid_sync_error: str | None = None
        window_start = None if all_years_enabled else start_year
        window_end = None if all_years_enabled else end_year

        in_window_citations = [citation for citation in citations if citation_in_window(citation, window_start, window_end)]
        in_window_pmids = {citation.pmid for citation in in_window_citations}
        excluded_out_of_window = 0 if all_years_enabled else (len(citations) - len(in_window_citations))
        excluded_pmids: set[str] = {citation.pmid for citation in citations if citation.pmid not in in_window_pmids}
        excluded_reasons: dict[str, str] = {
            citation.pmid: f"Outside year window {_year_window_label(window_start, window_end)}"
            for citation in citations
            if citation.pmid not in in_window_pmids
        }
        excluded_by_type = 0
        disambiguation_citations = in_window_citations
        if excluded_type_terms:
            type_filtered: list[Citation] = []
            for citation in in_window_citations:
                is_excluded, matched_types = citation_matches_excluded_type(citation, excluded_type_terms)
                if is_excluded:
                    excluded_by_type += 1
                    citation.notes.append(f"Excluded by publication type filter: {', '.join(matched_types)}")
                    excluded_pmids.add(citation.pmid)
                    excluded_reasons[citation.pmid] = f"Excluded by publication type: {', '.join(matched_types)}"
                    continue
                type_filtered.append(citation)
            disambiguation_citations = type_filtered
        logger.info(
            "Window/type filters before disambiguation: in_window=%d, excluded_out_of_window=%d, excluded_by_type=%d",
            len(disambiguation_citations),
            excluded_out_of_window,
            excluded_by_type,
        )

        clusters, eligible_matches = build_clusters(disambiguation_citations, target, target_orcid=target_orcid)
        _, all_matches = build_clusters(citations, target, target_orcid=target_orcid)

        if target_orcid:
            matched_affiliations_by_pmid: dict[str, list[str]] = {}
            for match in all_matches:
                affiliation = (match.author.affiliation or "").strip()
                if not affiliation:
                    continue
                matched_affiliations_by_pmid.setdefault(match.pmid, []).append(affiliation)
            if matched_affiliations_by_pmid:
                try:
                    orcid_affiliation_matches = orcid_client.match_affiliations(target_orcid, matched_affiliations_by_pmid)
                    logger.info(
                        "ORCiD affiliation cross-check completed: matched %d/%d citation(s)",
                        len(orcid_affiliation_matches),
                        len(matched_affiliations_by_pmid),
                    )
                except Exception as exc:  # noqa: BLE001
                    logger.warning("ORCiD affiliation cross-check failed for %s: %s", target_orcid, exc)
                    orcid_sync_warnings.append("affiliation cross-check unavailable")

        if orcid_sync_warnings:
            orcid_sync_error = "ORCiD " + "; ".join(orcid_sync_warnings)
            errors.append(orcid_sync_error)

        citation_by_pmid = {citation.pmid: citation for citation in disambiguation_citations}
        cluster_label_by_id = {cluster.cluster_id: cluster.label for cluster in clusters}
        cluster_citation_rows = _build_cluster_citation_rows(
            citations=disambiguation_citations,
            matches=eligible_matches,
            orcid_identifier_matches=orcid_identifier_matches,
            orcid_affiliation_matches=orcid_affiliation_matches,
            target_orcid=target_orcid,
        )
        cluster_orcid_affiliation_matches = _build_cluster_orcid_affiliation_matches(cluster_citation_rows)
        cluster_affiliation_display_rows = _build_cluster_affiliation_display_rows(
            clusters=clusters,
            cluster_orcid_affiliation_matches=cluster_orcid_affiliation_matches,
            target_orcid=target_orcid,
        )
        missing_affiliation_rows: list[dict[str, object]] = []
        for match in eligible_matches:
            matched_citation = citation_by_pmid.get(match.pmid)
            if matched_citation is None:
                continue
            affiliation = match.author.affiliation or ""
            reason = _missing_affiliation_reason(matched_citation, affiliation)
            row: dict[str, object] = {
                "cluster_label": cluster_label_by_id.get(match.cluster_id, match.cluster_id),
                "pmid": matched_citation.pmid,
                "pmid_link": f"https://pubmed.ncbi.nlm.nih.gov/{matched_citation.pmid}/",
                "pmcid": matched_citation.pmcid,
                "pmcid_link": f"https://pmc.ncbi.nlm.nih.gov/articles/{matched_citation.pmcid}/" if matched_citation.pmcid else None,
                "year": matched_citation.print_year,
                "source_for_roles": matched_citation.source_for_roles,
                "affiliation": affiliation or "(no affiliation listed)",
                "missing_reason": reason,
                "notes": "; ".join(matched_citation.notes),
                "source_badges": _citation_source_badges(matched_citation),
                **_orcid_row_fields(
                    pmid=matched_citation.pmid,
                    target_orcid=target_orcid,
                    orcid_identifier_matches=orcid_identifier_matches,
                    orcid_affiliation_matches=orcid_affiliation_matches,
                ),
            }
            if reason:
                missing_affiliation_rows.append(row)
        missing_affiliation_rows.sort(key=_dict_row_sort_key)
        excluded_citation_rows = _build_excluded_citation_rows(
            citations=citations,
            excluded_pmids=excluded_pmids,
            excluded_reasons=excluded_reasons,
            matches=all_matches,
            orcid_identifier_matches=orcid_identifier_matches,
            orcid_affiliation_matches=orcid_affiliation_matches,
            target_orcid=target_orcid,
        )
        (
            excluded_citation_rows_by_type,
            excluded_citation_rows_by_window,
            excluded_citation_rows_other,
        ) = _partition_excluded_citation_rows(excluded_citation_rows)
        logger.info(
            "Disambiguation computed: citations=%d, matches=%d, clusters=%d, errors=%d",
            len(citations),
            len(all_matches),
            len(clusters),
            len(errors),
        )

        run_id = uuid.uuid4().hex
        _save_run(
            run_id,
            RunState(
                author_name=author_name,
                target_orcid=target_orcid,
                start_year=window_start,
                end_year=window_end,
                excluded_type_terms=excluded_type_terms,
                citations=citations,
                clusters=clusters,
                matches=all_matches,
                excluded_pmids=excluded_pmids,
                excluded_reasons=excluded_reasons,
                orcid_identifier_matches=orcid_identifier_matches,
                orcid_affiliation_matches=orcid_affiliation_matches,
                orcid_sync_error=orcid_sync_error,
            ),
        )

        _update_usage_safe(
            usage_record_id,
            run_id=run_id,
            status=("completed_with_errors" if errors else "completed"),
            error_message="; ".join(errors),
            citation_count=len(citations),
            error_count=len(errors),
        )

        return templates.TemplateResponse(
            request,
            "disambiguate.html",
            {
                "run_id": run_id,
                "author_name": author_name,
                "target_orcid": target_orcid,
                "start_year": window_start,
                "end_year": window_end,
                "year_window_label": _year_window_label(window_start, window_end),
                "clusters": clusters,
                "cluster_citation_rows": cluster_citation_rows,
                "cluster_affiliation_display_rows": cluster_affiliation_display_rows,
                "errors": errors,
                "missing_affiliation_rows": missing_affiliation_rows,
                "excluded_citation_rows": excluded_citation_rows,
                "excluded_citation_rows_by_type": excluded_citation_rows_by_type,
                "excluded_citation_rows_by_window": excluded_citation_rows_by_window,
                "excluded_citation_rows_other": excluded_citation_rows_other,
                "selected_excluded_pmids": [],
                "excluded_out_of_window": excluded_out_of_window,
                "excluded_by_type": excluded_by_type,
                "extra_excluded_types": extra_excluded_types,
                "orcid_sync_error": orcid_sync_error,
            },
        )
    except Exception as exc:  # noqa: BLE001
        _update_usage_safe(
            usage_record_id,
            status="failed",
            error_message=str(exc),
        )
        raise
    finally:
        DISAMBIGUATION_SEMAPHORE.release()


@app.post("/citation-select", response_class=HTMLResponse)
def citation_select(
    request: Request,
    run_id: str = Form(...),
    selected_cluster_ids: list[str] = Form(default=[]),
    selected_excluded_pmids: list[str] = Form(default=[]),
) -> HTMLResponse:
    run = _load_run(run_id)
    selected = set(selected_cluster_ids)
    forced_excluded_pmids = set(selected_excluded_pmids)
    if not selected:
        eligible_citations = [citation for citation in run.citations if citation.pmid not in run.excluded_pmids]
        eligible_matches = [match for match in run.matches if match.pmid not in run.excluded_pmids]
        cluster_citation_rows = _build_cluster_citation_rows(
            citations=eligible_citations,
            matches=eligible_matches,
            orcid_identifier_matches=run.orcid_identifier_matches,
            orcid_affiliation_matches=run.orcid_affiliation_matches,
            target_orcid=run.target_orcid,
        )
        cluster_orcid_affiliation_matches = _build_cluster_orcid_affiliation_matches(cluster_citation_rows)
        cluster_affiliation_display_rows = _build_cluster_affiliation_display_rows(
            clusters=run.clusters,
            cluster_orcid_affiliation_matches=cluster_orcid_affiliation_matches,
            target_orcid=run.target_orcid,
        )
        excluded_citation_rows = _build_excluded_citation_rows(
            citations=run.citations,
            excluded_pmids=run.excluded_pmids,
            excluded_reasons=run.excluded_reasons,
            matches=run.matches,
            orcid_identifier_matches=run.orcid_identifier_matches,
            orcid_affiliation_matches=run.orcid_affiliation_matches,
            target_orcid=run.target_orcid,
        )
        (
            excluded_citation_rows_by_type,
            excluded_citation_rows_by_window,
            excluded_citation_rows_other,
        ) = _partition_excluded_citation_rows(excluded_citation_rows)
        return templates.TemplateResponse(
            request,
            "disambiguate.html",
            {
                "run_id": run_id,
                "author_name": run.author_name,
                "target_orcid": run.target_orcid,
                "start_year": run.start_year,
                "end_year": run.end_year,
                "year_window_label": _year_window_label(run.start_year, run.end_year),
                "clusters": run.clusters,
                "cluster_citation_rows": cluster_citation_rows,
                "cluster_affiliation_display_rows": cluster_affiliation_display_rows,
                "errors": ["Select at least one author identity cluster."],
                "missing_affiliation_rows": [],
                "excluded_citation_rows": excluded_citation_rows,
                "excluded_citation_rows_by_type": excluded_citation_rows_by_type,
                "excluded_citation_rows_by_window": excluded_citation_rows_by_window,
                "excluded_citation_rows_other": excluded_citation_rows_other,
                "selected_excluded_pmids": sorted(forced_excluded_pmids),
                "excluded_out_of_window": 0,
                "excluded_by_type": 0,
                "orcid_sync_error": run.orcid_sync_error,
            },
            status_code=400,
        )

    cluster_label_by_id = {cluster.cluster_id: cluster.label for cluster in run.clusters}
    selected_matches = _filter_matches_for_analysis(
        run=run,
        selected_cluster_ids=selected,
        selected_pmids=set(),
        pmid_filter_enabled=0,
        forced_excluded_pmids=forced_excluded_pmids,
    )
    citation_selection_rows = _build_citation_selection_rows(
        citations=run.citations,
        matches=selected_matches,
        all_matches=run.matches,
        cluster_label_by_id=cluster_label_by_id,
        start_year=run.start_year,
        end_year=run.end_year,
        orcid_identifier_matches=run.orcid_identifier_matches,
        orcid_affiliation_matches=run.orcid_affiliation_matches,
        target_orcid=run.target_orcid,
    )
    needs_review = any(bool(row["needs_review"]) for row in citation_selection_rows)
    logger.info(
        "Citation-selection gate: run_id=%s, selected_clusters=%d, matched_citations=%d, needs_review=%s",
        run_id,
        len(selected),
        len(citation_selection_rows),
        needs_review,
    )

    if not needs_review:
        analysis = analyze_citations(
            citations=run.citations,
            matches=selected_matches,
            selected_cluster_ids=selected,
            start_year=run.start_year,
            end_year=run.end_year,
            forced_include_pmids=forced_excluded_pmids,
            excluded_pmids=run.excluded_pmids,
        )
        out_of_window_rows = _build_out_of_window_rows(
            citations=run.citations,
            matches=selected_matches,
            start_year=run.start_year,
            end_year=run.end_year,
            forced_include_pmids=forced_excluded_pmids,
            orcid_identifier_matches=run.orcid_identifier_matches,
            orcid_affiliation_matches=run.orcid_affiliation_matches,
            target_orcid=run.target_orcid,
        )
        return _render_analysis_result(
            request,
            run=run,
            run_id=run_id,
            selected_cluster_ids=selected,
            selected_pmids=set(),
            selected_excluded_pmids=forced_excluded_pmids,
            pmid_filter_enabled=0,
            analysis=analysis,
            start_year=run.start_year,
            end_year=run.end_year,
            out_of_window_rows=out_of_window_rows,
        )

    return templates.TemplateResponse(
        request,
        "citation_select.html",
        {
            "run_id": run_id,
            "author_name": run.author_name,
            "target_orcid": run.target_orcid,
            "orcid_sync_error": run.orcid_sync_error,
            "start_year": run.start_year,
            "end_year": run.end_year,
            "year_window_label": _year_window_label(run.start_year, run.end_year),
            "selected_cluster_ids": sorted(selected),
            "selected_excluded_pmids": sorted(forced_excluded_pmids),
            "citation_selection_rows": [row for row in citation_selection_rows if bool(row["needs_review"])],
            "auto_included_pmids": sorted(
                {
                    str(row["pmid"])
                    for row in citation_selection_rows
                    if not bool(row["needs_review"])
                }
            ),
        },
    )
@app.post("/analyze", response_class=HTMLResponse)
def analyze(
    request: Request,
    run_id: str = Form(...),
    selected_cluster_ids: list[str] = Form(default=[]),
    selected_pmids: list[str] = Form(default=[]),
    selected_excluded_pmids: list[str] = Form(default=[]),
    pmid_filter_enabled: int = Form(default=0),
) -> HTMLResponse:
    run = _load_run(run_id)
    selected = set(selected_cluster_ids)
    selected_pmid_set = set(selected_pmids)
    forced_excluded_pmids = set(selected_excluded_pmids)
    logger.info(
        "Analyze request: run_id=%s, selected_clusters=%d, selected_pmids=%d, forced_excluded=%d, pmid_filter_enabled=%d",
        run_id,
        len(selected),
        len(selected_pmid_set),
        len(forced_excluded_pmids),
        pmid_filter_enabled,
    )

    filtered_matches = _filter_matches_for_analysis(
        run=run,
        selected_cluster_ids=selected,
        selected_pmids=selected_pmid_set,
        pmid_filter_enabled=pmid_filter_enabled,
        forced_excluded_pmids=forced_excluded_pmids,
    )

    analysis = analyze_citations(
        citations=run.citations,
        matches=filtered_matches,
        selected_cluster_ids=selected,
        start_year=run.start_year,
        end_year=run.end_year,
        forced_include_pmids=forced_excluded_pmids,
        excluded_pmids=run.excluded_pmids,
    )
    out_of_window_rows = _build_out_of_window_rows(
        citations=run.citations,
        matches=filtered_matches,
        start_year=run.start_year,
        end_year=run.end_year,
        forced_include_pmids=forced_excluded_pmids,
        orcid_identifier_matches=run.orcid_identifier_matches,
        orcid_affiliation_matches=run.orcid_affiliation_matches,
        target_orcid=run.target_orcid,
    )
    return _render_analysis_result(
        request,
        run=run,
        run_id=run_id,
        selected_cluster_ids=selected,
        selected_pmids=selected_pmid_set,
        selected_excluded_pmids=forced_excluded_pmids,
        pmid_filter_enabled=pmid_filter_enabled,
        analysis=analysis,
        start_year=run.start_year,
        end_year=run.end_year,
        out_of_window_rows=out_of_window_rows,
    )


@app.post("/finalize", response_class=HTMLResponse)
def finalize(
    request: Request,
    run_id: str = Form(...),
    selected_cluster_ids: list[str] = Form(default=[]),
    selected_pmids: list[str] = Form(default=[]),
    selected_excluded_pmids: list[str] = Form(default=[]),
    pmid_filter_enabled: int = Form(default=0),
    accepted_row_indices: list[int] = Form(default=[]),
    correction_entries: list[str] = Form(default=[]),
) -> HTMLResponse:
    run = _load_run(run_id)
    selected = set(selected_cluster_ids)
    selected_pmid_set = set(selected_pmids)
    forced_excluded_pmids = set(selected_excluded_pmids)
    accepted = set(accepted_row_indices)
    logger.info(
        "Finalize request: run_id=%s, selected_clusters=%d, selected_pmids=%d, forced_excluded=%d, pmid_filter_enabled=%d, accepted_uncertain=%d, corrections=%d",
        run_id,
        len(selected),
        len(selected_pmid_set),
        len(forced_excluded_pmids),
        pmid_filter_enabled,
        len(accepted),
        len(correction_entries),
    )

    filtered_matches = _filter_matches_for_analysis(
        run=run,
        selected_cluster_ids=selected,
        selected_pmids=selected_pmid_set,
        pmid_filter_enabled=pmid_filter_enabled,
        forced_excluded_pmids=forced_excluded_pmids,
    )

    analysis = analyze_citations(
        citations=run.citations,
        matches=filtered_matches,
        selected_cluster_ids=selected,
        start_year=run.start_year,
        end_year=run.end_year,
        forced_include_pmids=forced_excluded_pmids,
        excluded_pmids=run.excluded_pmids,
    )

    _apply_uncertain_selection_and_corrections(
        analysis=analysis,
        accepted_row_indices=accepted,
        correction_entries=correction_entries,
    )
    out_of_window_rows = _build_out_of_window_rows(
        citations=run.citations,
        matches=filtered_matches,
        start_year=run.start_year,
        end_year=run.end_year,
        forced_include_pmids=forced_excluded_pmids,
        orcid_identifier_matches=run.orcid_identifier_matches,
        orcid_affiliation_matches=run.orcid_affiliation_matches,
        target_orcid=run.target_orcid,
    )
    return _render_analysis_result(
        request,
        run=run,
        run_id=run_id,
        selected_cluster_ids=selected,
        selected_pmids=selected_pmid_set,
        selected_excluded_pmids=forced_excluded_pmids,
        pmid_filter_enabled=pmid_filter_enabled,
        analysis=analysis,
        start_year=run.start_year,
        end_year=run.end_year,
        force_report=True,
        out_of_window_rows=out_of_window_rows,
    )


@app.post("/report-correct", response_class=HTMLResponse)
def report_correct(
    request: Request,
    run_id: str = Form(...),
    selected_cluster_ids: list[str] = Form(default=[]),
    selected_pmids: list[str] = Form(default=[]),
    selected_excluded_pmids: list[str] = Form(default=[]),
    pmid_filter_enabled: int = Form(default=0),
    accepted_uncertain_pmids: list[str] = Form(default=[]),
    correction_entries: list[str] = Form(default=[]),
    download: str | None = Form(default=None),
) -> Response:
    run = _load_run(run_id)
    selected = set(selected_cluster_ids)
    selected_pmid_set = set(selected_pmids)
    forced_excluded_pmids = set(selected_excluded_pmids)
    accepted_uncertain_set = set(accepted_uncertain_pmids)
    logger.info(
        "Report-correct request: run_id=%s, selected_clusters=%d, selected_pmids=%d, forced_excluded=%d, corrections=%d",
        run_id,
        len(selected),
        len(selected_pmid_set),
        len(forced_excluded_pmids),
        len(correction_entries),
    )

    analysis, filtered_matches = _rebuild_analysis_with_overrides(
        run=run,
        selected_cluster_ids=selected,
        selected_pmids=selected_pmid_set,
        selected_excluded_pmids=forced_excluded_pmids,
        pmid_filter_enabled=pmid_filter_enabled,
        accepted_uncertain_pmids=accepted_uncertain_set,
        correction_entries=correction_entries,
    )
    if (download or "").strip().lower() == "xlsx":
        return _xlsx_response_for_analysis(run=run, analysis=analysis)

    out_of_window_rows = _build_out_of_window_rows(
        citations=run.citations,
        matches=filtered_matches,
        start_year=run.start_year,
        end_year=run.end_year,
        forced_include_pmids=forced_excluded_pmids,
        orcid_identifier_matches=run.orcid_identifier_matches,
        orcid_affiliation_matches=run.orcid_affiliation_matches,
        target_orcid=run.target_orcid,
    )
    return _render_analysis_result(
        request,
        run=run,
        run_id=run_id,
        selected_cluster_ids=selected,
        selected_pmids=selected_pmid_set,
        selected_excluded_pmids=forced_excluded_pmids,
        pmid_filter_enabled=pmid_filter_enabled,
        analysis=analysis,
        start_year=run.start_year,
        end_year=run.end_year,
        force_report=True,
        out_of_window_rows=out_of_window_rows,
    )
